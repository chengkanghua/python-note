'''
基于csv格式实现 用户的注册 & 登录认证。详细需求如下：

- 用户注册时，新注册用户要写入文件csv文件中，输入Q或q则退出。
- 用户登录时，逐行读取csv文件中的用户信息并进行校验。
- 提示：文件路径须使用os模块构造的绝对路径的方式。
'''

'''
import os

# 定位文件
base_dir = os.path.dirname(os.path.abspath(__file__))
db_file_path = os.path.join(base_dir,'files/db1.csv')

# 用户注册
while True:
    choice = input('enter registry Y/y， quit N/n: ')
    if not choice.upper() in {'Y','N'}:
        print('error: repeat')
        continue
    if choice.upper() == 'N':
        break
    with open(db_file_path,mode='a',encoding='utf-8') as f:
        while True:
            username = input('please username: Q/q quit:')
            if username.upper() == 'Q':
                break
            password = input('please passwrod:')
            f.write(f'{username},{password}\n')
            f.flush()
    break

# 用户登录
username = input('wecome login  please username: ')
password = input('please password: ')

if not os.path.exists(db_file_path):
    print('error: db-file not find ')
else:
    with open(db_file_path,mode='r',encoding='utf-8') as f_read:
        for line in f_read:
            uname,pwd = line.strip().split(',')
            if uname == username and pwd == password:
                print('login success：')
                break
        else:
            print('error：username or password wroing')
'''

# 补充代码：实现去网上获取指定地区的天气信息，并写入到Excel中。
'''
import os
import requests
from xml.etree import cElementTree as ET
from openpyxl import workbook

#  文件定位
base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir,'files/p1.xlsx')
print(file_path)
# 创建excel
wb = workbook.Workbook()
del wb['Sheet']

while True:
    city = input("请输入城市（Q/q退出）：")
    if city.upper() == "Q":
        break
    url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
    res = requests.get(url=url)
    # print(res.text)

    # 1.提取XML格式中的数据
    root = ET.XML(res.text)
    # 2.为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。
    sheet = wb.create_sheet(city)
    # sheet = wb[city]
    for index_row,val in enumerate(root,1):
        # print(index,val.text)
        sheet.cell(index_row,1).value = val.text
        # cell = sheet.cell(index_row,1)
        # cell.value = val.text
wb.save(file_path)
'''