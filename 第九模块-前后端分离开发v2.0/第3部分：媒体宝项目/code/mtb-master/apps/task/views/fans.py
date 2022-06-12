import os
import requests
from utils.token import get_authorizer_access_token
from django.conf import settings
from utils.ext.mixins import MtbCreateModelMixin, MtbDestroyModelMixin, MtbUpdateModelMixin, MtbListModelMixin
from rest_framework.viewsets import GenericViewSet
from utils.ext.filters import SelfFilterBackend
from rest_framework.filters import BaseFilterBackend
from rest_framework.views import APIView
from rest_framework.response import Response
from utils import return_code
from django.shortcuts import HttpResponse

from ..serializers.fans import FansSerializer
from .. import models

from utils.ext.page import MtbPageNumberPagination
from utils.ext.mixins import MtbPageNumberListModelMixin
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import FilterSet, filters


class FansFilterSet(FilterSet):
    name = filters.CharFilter(field_name='nick_name', lookup_expr='contains')
    activity = filters.NumberFilter(field_name='activity_id')
    promo = filters.NumberFilter(field_name='origin_open_id')
    black = filters.NumberFilter(field_name='black')  # 0/1

    class Meta:
        model = models.TakePartIn
        fields = ["name", "activity", "promo", "black"]


class SelfPublicFilterBackend(BaseFilterBackend):
    def filter_queryset(self, request, queryset, view):
        return queryset.filter(public_number__mtb_user_id=request.user.user_id)


class FansView(MtbPageNumberListModelMixin, GenericViewSet):
    filter_backends = [SelfPublicFilterBackend, DjangoFilterBackend]

    filterset_class = FansFilterSet
    pagination_class = MtbPageNumberPagination

    queryset = models.TakePartIn.objects.all().order_by('-id')
    serializer_class = FansSerializer


class ToBlackView(APIView):
    def post(self, request, *args, **kwargs):
        id_list = request.data.get('id_list', [])
        models.TakePartIn.objects.filter(public_number__mtb_user=request.user.user_id, id__in=id_list).update(black=1)
        return Response({"code": return_code.SUCCESS})


class OutBlackView(APIView):
    def post(self, request, *args, **kwargs):
        id_list = request.data.get('id_list', [])
        models.TakePartIn.objects.filter(public_number__mtb_user=request.user.user_id, id__in=id_list).update(black=0)
        return Response({"code": return_code.SUCCESS})


from utils.ext.auth import ParamsJwtTokenAuthentication


class ExportBlackView(APIView):
    authentication_classes = [ParamsJwtTokenAuthentication, ]

    def get(self, request, *args, **kwargs):

        # 1.根据参数读取数据，写入Excel
        search_map = {"name": "name__contains", "activity": "activity_id", 'origin_open_id': "origin_open_id"}
        condition = {"public_number__mtb_user_id": request.user.user_id}
        for k, v in search_map.items():
            data = request.query_params.get(k)
            if not data:
                continue
            condition[v] = data
        queryset = models.TakePartIn.objects.filter(**condition)

        # 2.读取数据写入Excel    (pip install openpyxl)
        import io
        from openpyxl import workbook
        from django.utils.encoding import escape_uri_path
        wb = workbook.Workbook()
        sheet = wb.worksheets[0]
        next_row_index = 1
        for obj in queryset:
            sheet.cell(next_row_index, 1).value = obj.nick_name
            sheet.cell(next_row_index, 2).value = obj.open_id
            sheet.cell(next_row_index, 3).value = obj.get_looking_display()
            sheet.cell(next_row_index, 4).value = obj.public_number.nick_name
            next_row_index += 1
        stream = io.BytesIO()
        wb.save(stream)
        # 3.将内容导出
        response = HttpResponse(stream.getbuffer(), content_type="application/octet-stream")
        response['Content-Disposition'] = "attachment; filename={};".format(escape_uri_path("temp.xlsx"))
        return response
