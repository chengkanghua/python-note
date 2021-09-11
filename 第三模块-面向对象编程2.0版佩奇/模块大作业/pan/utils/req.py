import os,time,datetime,struct
from openpyxl import load_workbook
from config import settings



def send_data(conn: object,content: str):
    content = content.encode('utf-8')
    header = struct.pack('i',len(content))
    conn.sendall(header)
    conn.sendall(content)

def recv_data(conn: object,chunk_size = 1024):
    #获取头部数据  数据长度
    has_recv_size = 0
    bytes_list = []
    while has_recv_size < 4:
        chunk = conn.recv(4 - has_recv_size)
        bytes_list.append(chunk)
        has_recv_size += len(chunk)
    data = b''.join(bytes_list)
    data_length = struct.unpack('i',data)[0]

    #获取数据
    has_recv_size = 0
    data_list = []
    while data_length > has_recv_size:
        size = chunk_size if (data_length - has_recv_size) > chunk_size else data_length - has_recv_size
        chunk =conn.recv(size)
        data_list.append(chunk)
        has_recv_size += len(chunk)
    data = b''.join(data_list)
    return data

def recv_save_file(conn: object,save_file_path: str ,chunk_size=1024):
    #接受头部数据 数据长度
    has_recv_size = 0
    bytes_list = []
    while has_recv_size < 4:
        data = conn.recv(4 - has_recv_size)
        bytes_list.append(data)
        has_recv_size += len(data)
    data = b''.join(bytes_list)
    data_length = struct.unpack('i',data)[0]

    #接受数据存入硬盘
    has_recv_size = 0
    with open(save_file_path,mode='wb') as f:
        while has_recv_size < data_length:
            size = chunk_size if (data_length - has_recv_size) > chunk_size else data_length - has_recv_size
            chunk = conn.recv(size)
            f.write(chunk)
            f.flush()
            has_recv_size += len(chunk)


def send_file_by_seek(conn,file_size,file_path,seek=0):
    #发送头部数据 长度
    header = struct.pack('i',file_size)
    conn.sendall(header)

    has_send_size = 0
    with open(file_path,mode='rb') as f:
        if seek:
            f.seek(seek)
            has_send_size += seek
        while file_size > has_send_size:
            chunk = f.read(2048)
            conn.sendall(chunk)
            has_send_size += len(chunk)

def login(username,password):
    """ 用户登录，读取excel文件，进行用户登录 """
    wb = load_workbook(settings.DB_FILE_PATH)  # 载入excel 文件
    sheet = wb.worksheets[0]  # 第一个sheet

    success = False
    for row in sheet.iter_rows(2):  # 从第二行开始遍历
        if username == row[0].value and password == row[1].value:
            success = True
            break
    return success

def is_exists_user(username):
    wb = load_workbook(settings.DB_FILE_PATH)
    sheet = wb.worksheets[0]
    exists = False
    for row in sheet.iter_rows(2):
        if username == row[0].value:
            exists = True
    return exists

def register_user(username,password):
    wb = load_workbook(settings.DB_FILE_PATH)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    data_list = [username,password,datetime.datetime.now().strftime('%Y-%m-%d')]
    for i,item in enumerate(data_list,1):
        cell = sheet.cell(max_row+1,i)
        cell.value = item
    wb.save(settings.DB_FILE_PATH)
    user_folder = os.path.join(settings.USER_FOLDER_PATH,username)
    os.makedirs(user_folder)
    return True

