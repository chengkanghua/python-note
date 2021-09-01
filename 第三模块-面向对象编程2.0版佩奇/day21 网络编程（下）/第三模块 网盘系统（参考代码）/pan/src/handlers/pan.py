import re
import os
import json
import time
import datetime

from openpyxl import load_workbook

from utils import req
from config import settings


class PanHandler(object):

    def __init__(self, conn):
        # 客户端的socket连接
        self.conn = conn
        self.username = None

    @property
    def home_path(self):
        return os.path.join(settings.USER_FOLDER_PATH, self.username)

    def send_json_data(self, **kwargs):
        # kwargs={"status": False, 'error': "用户名已存在"}
        req.send_data(self.conn, json.dumps(kwargs))

    def recv_save_file(self, target_file_path):
        req.recv_save_file(self.conn, target_file_path)

    def send_file_by_seek(self, file_size, file_path, seek=0):
        req.send_file_by_seek(self.conn, file_size, file_path, seek)

    def login(self, username, password):
        """ 用户登录，读取excel文件，进行用户登录 """
        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        success = False
        for row in sheet.iter_rows(2):
            if username == row[0].value and password == row[1].value:
                success = True
                break

        if success:
            # req.send_data(self.conn, json.dumps({"status": True, 'data': "登录成功"}))
            self.send_json_data(status=True, data="登录成功")
            self.username = username
        else:
            # req.send_data(self.conn, json.dumps({"status": False, 'error': "登录失败"}))
            self.send_json_data(status=False, error="登录失败")

    def register(self, username, password):
        """ 用户注册， 用户名和密码写入到excel中（已存在则不再注册） """

        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        # 监测用户是否已经存在
        exists = False
        for row in sheet.iter_rows(2):  # for row in sheet.rows，是从头开始读取（含表头）
            if username == row[0].value:
                exists = True
                break
        if exists:
            # 给客户端回复：用户名已存在
            # self.conn.sendall(b"....") -> 粘包
            # req.send_data(self.conn, json.dumps({"status": False, 'error': "用户名已存在"}))
            self.send_json_data(status=False, error="用户名已存在")
            return

        # 注册用户写入excel
        max_row = sheet.max_row
        data_list = [username, password, datetime.datetime.now().strftime("%Y-%m-%d")]
        for i, item in enumerate(data_list, 1):
            cell = sheet.cell(max_row + 1, i)
            cell.value = item
        wb.save(settings.DB_FILE_PATH)

        # 创建用户目录
        user_folder = os.path.join(settings.USER_FOLDER_PATH, username)
        os.makedirs(user_folder)

        # 回复消息
        # req.send_data(self.conn, json.dumps({"status": True, 'data': "注册成功"}))
        self.send_json_data(status=True, data="注册成功")

    def ls(self, folder_path=None):
        """ 查看当前用户目录下所有的文件
            1. folder_path=None，查看用户根目录
            2. folder_path不为空，查看  用户目录/folder_path 中的文件
        """
        if not self.username:
            self.send_json_data(status=False, error="登录后才能查看")
            return

        if not folder_path:
            # 自己的根目录： files目录 + 用户名
            data = "\n".join(os.listdir(self.home_path))
            self.send_json_data(status=True, data=data)
            return

        target_folder = os.path.join(self.home_path, folder_path)

        if not os.path.exists(target_folder):
            self.send_json_data(status=False, error="路径不存在")
            return
        if not os.path.isdir(target_folder):
            self.send_json_data(status=False, error="文件夹不存在")
            return

        data = "\n".join(os.listdir(target_folder))
        self.send_json_data(status=True, data=data)

    def upload(self, file_path):
        """ 上传文件，直接覆盖"""
        # 用户未登录
        if not self.username:
            self.send_json_data(status=False, error="登录后才能查看")
            return

        target_file_path = os.path.join(self.home_path, file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.makedirs(folder)

        self.send_json_data(status=True, data="开始上传")

        # *** 很关键，如果是基于 select + 非阻塞的模式下，客户度未发送数据，直接去recv，会出现blocking error的错误 ***
        # 解决方案：1.改为阻塞模式； 2.使用time.sleep等一会，等有数据了再去读取。
        # time.sleep(2)

        # 接收上传的文件
        self.recv_save_file(target_file_path)

    def download(self, file_path, seek=0):
        """ 下载文件，支持断点续传（客户端本地已有文件）
            seek=None，从头开始下载；
            seek=1000，从1000字节开始下载（续传）
        """
        # 用户未登录
        if not self.username:
            # req.send_data(self.conn, json.dumps({"status": False, "error": "登录成功后才能上传"}))
            self.send_json_data(status=False, error="登录成功后才能上传")
            return

        # 文件不存在
        target_file_path = os.path.join(self.home_path, file_path)
        if not os.path.exists(target_file_path):
            # req.send_data(self.conn, json.dumps({"status": False, "error": "文件{}不存在".format(file_path)}))
            self.send_json_data(status=False, error="文件{}不存在".format(file_path))
            return

        # 获取文件大小并返回
        # req.send_data(self.conn, json.dumps({"status": True, "data": "开始下载"}))
        self.send_json_data(status=True, data="开始下载")

        # 发送文件
        seek = int(seek)
        total_size = os.stat(target_file_path).st_size
        req.send_file_by_seek(self.conn, total_size - seek, target_file_path, seek)

    def execute(self):
        """
        每次客户端发来请求，触发此方法。
        :return: False，关闭连接；True，继续处理请求
        """
        conn = self.conn

        # 登录/注册/查看目录/上传文件/下载文件

        # 1. 获取数据包
        cmd = req.recv_data(conn).decode('utf-8')
        if cmd.upper() == "Q":
            print("客户端退出")
            return False
        """
        login wupeiqi 123
        register wupeiqi 123
        ls   ls xxx（登录成功之后）
        upload v4.py（登录成功之后）
        download v1.txt（登录成功之后）
        """
        method_map = {
            "login": self.login,
            "register": self.register,
            "ls": self.ls,
            "upload": self.upload,
            "download": self.download,
        }

        # "register  root   666"
        # "login  alex   123"
        # "ls"      [ls,]
        # "ls xxx"  [ls,xxx]
        # upload v2.py
        cmd, *args = re.split(r"\s+", cmd)  # [register,root,123] /  [login,root,123]
        method = method_map[cmd]

        # 2. 执行相关方法
        method(*args)

        return True
