import os
import re
import select
import socket
import datetime
import config as config
import utils.req as req
from openpyxl import load_workbook


class server():
    def __init__(self):
        self.user = None

    def run(self):
        server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server.bind((config.HOST, config.PORT))
        server.listen(5)

        inputs = [server, ]
        while True:
            r, w, e = select.select(inputs, [], [], 0.05)
            for sock in r:
                if sock == server:
                    print("新客户端连接")
                    conn, addr = server.accept()
                    inputs.append(conn)
                    continue

            conn, addr = server.accept()
            while True:
                choice, *args = re.split(r"\s+", req.recv_data(conn).decode("utf-8"))
                if choice.upper() == "Q":
                    break
                print("选择", choice)

                method = {
                    "login": self.login,
                    "register": self.register,
                    "ls": self.ls,
                    "upload": self.upload,
                    "download": self.download,
                }

                if method[choice]:
                    method[choice](conn, *args)
            conn.close()

    def login(self, conn, *args):
        user, pwd = args
        wb = load_workbook(config.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        for row in sheet.rows:
            if user == row[0].value and pwd == row[1].value:
                req.send_data(conn, "登陆成功")
                self.user = user
                return
        req.send_data(conn, "登录失败")

    def register(self, conn, *args):
        user, pwd = args
        wb = load_workbook(config.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        # 判断用户是否已经存在
        exists = False
        for row in sheet.iter_rows(2):  # for row in sheet.rows，是从头开始读取（含表头）
            if user == row[0].value:
                exists = True
                break
        if exists:
            # 给客户端回复：用户名已存在
            req.send_data(conn, "用户名已存在")
            return

        # 注册用户写入excel
        max_row = sheet.max_row
        data_list = [user, pwd, datetime.datetime.now().strftime("%Y-%m-%d")]
        for i, item in enumerate(data_list, 1):
            cell = sheet.cell(max_row + 1, i)
            cell.value = item
        wb.save(config.DB_FILE_PATH)

        # 创建用户目录
        user_folder = os.path.join(config.USER_FOLDER_PATH, user)
        os.makedirs(user_folder)

        req.send_data(conn, "注册成功")

    def ls(self, conn, dir=None):
        if not self.user:
            req.send_data(conn, "请先登录在查看")
            return

        # 根目录
        if not dir:
            home_path = os.path.join(config.USER_FOLDER_PATH, self.user)
            data = "\n".join(os.listdir(home_path))
            req.send_data(conn, data)
            return

        file_path = os.path.join(config.USER_FOLDER_PATH, self.user, dir)
        # 判断查找目录是否存在
        if not os.path.exists(file_path):
            req.send_data(conn, "文件不存在")
            return

        data = os.listdir(file_path)
        if not data:
            req.send_data(conn, "空的文件夹")
            return

        data = "\n".join(data)
        req.send_data(conn, data)

    def upload(self, conn, remote_path):
        if not self.user:
            req.send_data(conn, "请先登录在查看")
            return

        remote_path = os.path.join(config.USER_FOLDER_PATH, self.user, remote_path)
        dir = os.path.dirname(remote_path)
        if not os.path.exists(dir):
            os.makedirs(dir)
        req.send_data(conn, "开始上传")

        req.recv_save_file(conn, remote_path)

    def download(self, conn, remote_path, seek=0):
        if not self.user:
            req.send_data(conn, "请先登录在查看")
            return
        remote_path = os.path.join(config.USER_FOLDER_PATH, self.user, remote_path)
        if not os.path.exists(remote_path):
            req.send_data(conn, "下载文件不存在")
            return
        seek = int(seek)
        if seek == os.stat(remote_path).st_size:
            req.send_data(conn, "已下载完成。")
            return
        req.send_data(conn, "开始下载")
        req.send_file_by_seek(conn, remote_path, seek)
